from openpyxl import Workbook,load_workbook
from matplotlib import pyplot as plt
from matplotlib.ticker import FuncFormatter
import matplotlib.gridspec as gridspec
# import os
# import numpy as np
# import pandas as pd


# 1. 用 openpyxl 读取 Excel 数据
# 打开 Excel 文件（注意：文件路径需正确，若同目录直接写文件名）
wb = Workbook()
book = load_workbook('20251118-MSD.xlsx',data_only = True)
sheet = book.active

# 2. 选择工作表（按表名或索引，索引从0开始）提取表头
headers = [book.sheetnames[0]]  
# print(headers[1:])##################################################################################################需要注释
# print(headers)############################################################################################需要注释

# 3. 初始化数据存储列表（时间列+各浓度MSD列）
time_data = []  # 时间数据（x轴）
msd_data = {}   # MSD数据（y轴），key=浓度，value=对应MSD值列表

for header in sheet.iter_rows(min_row=3,min_col=5,max_row = 3, values_only=True):  # 跳过时间列，初始化各浓度的MSD列表
    for i in header:
        msd_data[i] = []
# print("28行代码输出的msd_data:{}".format(msd_data))#######################################################################################################需要注释

# 4. 遍历行读取数据（从第5行第4列开始取数据，跳过表头）##怎么把元组区分开，并按顺序分别输入到字典的值中
for row in sheet.iter_rows(min_row=5,min_col=4, values_only=True):
    # print(row)############################################################################################需要注释
    # 读取时间（第4列），过滤空值
    if row[0] is not None:
        time_data.append(row[0])
        # 读取各浓度的MSD值（第4列及以后）
for row in sheet.iter_rows(min_row=5,min_col=5, values_only=True):
    for key, value in zip(msd_data.keys(),row):###########半自动根据数据的列数自己调整
                    msd_data[key].append(value)


# print("41行代码输出的time_data:{}".format(time_data))#################################################################################################需要注释
# print("42行代码输出的msd_data:{}".format(msd_data))###########################################################################################################需要注释

book.close()
print(f"成功读取数据：时间点数量={len(time_data)}，浓度类型={list(msd_data.keys())}")

# 4. 设置matplotlib样式（解决中文乱码+图表美观）
plt.rcParams["font.family"] = ["Times New Roman", "Microsoft YaHei"]  # 中文支持
plt.rcParams["axes.unicode_minus"] = False  # 负号正常显示
plt.rcParams["figure.facecolor"] = "white"  # 图表背景色

# 5. 绘制散点图（不同浓度用不同颜色/标记区分）
plt.figure(figsize=(12, 8))  # 图表大小（宽12，高8）

# 定义颜色和标记（区分不同浓度，避免混淆）
color_list = ["#FF6B6B", "#4ECDC4", "#45B7D1", "#FFA07A", "#98D8C8"]
marker_list = ["o", "s", "^", "D", "p"]

# 循环绘制各浓度的散点
for i, (concentration, msd_values) in enumerate(msd_data.items()):
    # print("62行代码：{}".format(msd_values))
    # print("63行代码: {}".format(time_data[:len(msd_values)]))
    # print("64行代码: {}".format(concentration))
    plt.scatter(
        time_data[:len(msd_values)],  # 确保时间与MSD数据长度一致
        msd_values,
        label=f"{concentration}",     # 图例标签（浓度）
        color=color_list[i % len(color_list)],  # 循环用色
        marker=marker_list[i % len(marker_list)],  # 循环用标记
        s=80,  # 点的大小
        alpha=0.8,  # 点的透明度（避免重叠遮挡）
        edgecolors="black",  # 点的边缘色（增强辨识度）
        linewidth=0.5  # 点的边缘线宽
    )

# 设置图表标签和标题
plt.title("不同ACN浓度下MSD随时间变化散点图", fontsize=14, fontweight="bold", pad=20)
plt.xlabel("时间 (ps)", fontsize=12, labelpad=10)
plt.ylabel("MSD (Å²)", fontsize=12, labelpad=10)

# 设置网格（辅助读取数据）
plt.grid(True, linestyle="--", alpha=0.5, color="#CCCCCC")

# 设置图例（放在右上角，避免遮挡数据）
plt.legend(
    title="ACN浓度", 
    title_fontsize=10, 
    fontsize=9, 
    loc="upper left", 
    bbox_to_anchor=(1, 1)  # 图例靠右侧显示
)

# 调整布局（防止标签/图例被截断）
plt.tight_layout()

plt.show()






